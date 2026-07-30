from __future__ import annotations

import csv
import io
import os
import threading
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

import duckdb
from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analysis import (
    DIMENSIONS,
    METRICS,
    AnalysisRequest,
    ExtractionRequest,
    analysis_metadata,
    extraction_columns,
    extraction_preview,
    extraction_rows,
    hierarchy_options,
    run_analysis,
)
from .csp import (
    CSP_DIMENSIONS,
    CSP_MEASURES,
    CspEvolutionRequest,
    CspExtractionRequest,
    CspOverviewRequest,
    csp_extraction_columns,
    csp_extraction_preview,
    csp_extraction_rows,
    csp_evolution,
    csp_metadata,
    csp_overview,
)
from .pathologies import (
    PATHOLOGY_DIMENSIONS,
    PATHOLOGY_MEASURES,
    PathologyExtractionRequest,
    PathologyOverviewRequest,
    pathology_extraction_columns,
    pathology_extraction_preview,
    pathology_extraction_rows,
    pathology_metadata,
    pathology_overview,
)
from .mortality import (
    MORTALITY_DIMENSIONS,
    MORTALITY_GROUPS,
    MORTALITY_MEASURES,
    MortalityExtractionRequest,
    MortalityOverviewRequest,
    mortality_extraction_columns,
    mortality_extraction_preview,
    mortality_extraction_rows,
    mortality_metadata,
    mortality_overview,
)
from .studio import (
    WorkbenchRequest,
    methodology,
    panorama,
    run_workbench,
    studio_metadata,
)


APP_DIR = Path(__file__).resolve().parent
WEBAPP_DIR = APP_DIR.parent.parent
PROJECT_DIR = WEBAPP_DIR.parent
DATA_DIR = Path(os.environ.get("DAMIR_DATA_DIR", PROJECT_DIR)).resolve()
CUBE_PATH = DATA_DIR / "cube_damir.parquet"
DELAYS_PATH = DATA_DIR / "cube_delais.parquet"
TRANSCO_PATH = DATA_DIR / "prs_nat_transco.csv"
PATHOLOGIES_PATH = Path(os.environ.get(
    "PATHOLOGIES_DATA_PATH", PROJECT_DIR / "Outil_patho" / "effectifs.parquet"
)).resolve()
CSP_DATA_ENV = os.environ.get("CSP_DATA_PATH")
CSP_DIR = (PROJECT_DIR / "CSP").resolve()
CSP_PATH = Path(CSP_DATA_ENV).resolve() if CSP_DATA_ENV else (CSP_DIR / "csp_core_2023.parquet").resolve()
CSP_GEOJSON_PATH = Path(os.environ.get(
    "CSP_GEOJSON_PATH", PROJECT_DIR / "Outil_patho" / "regions.geojson"
)).resolve()
MORTALITY_DATA_ENV = os.environ.get("MORTALITY_DATA_PATH")
MORTALITY_DIR = (PROJECT_DIR / "Mortalité").resolve()
MORTALITY_XLSX_PATH = MORTALITY_DIR / "Effectifs_cause_mortalite.xlsx"
MORTALITY_PARQUET_PATH = MORTALITY_DIR / "mortalite_core.parquet"
# The compact Parquet is the runtime source.  The workbook remains a safe
# fallback so a fresh checkout can still start before the conversion step.
MORTALITY_PATH = Path(MORTALITY_DATA_ENV).resolve() if MORTALITY_DATA_ENV else (
    MORTALITY_PARQUET_PATH if MORTALITY_PARQUET_PATH.exists() else MORTALITY_XLSX_PATH
).resolve()
FRONTEND_DIST = WEBAPP_DIR / "frontend" / "dist"
FRONTEND_ASSETS = FRONTEND_DIST / "assets"

REGIONS = {
    1: "Guadeloupe",
    2: "Martinique",
    3: "Guyane",
    4: "La Réunion",
    5: "DOM (ancien code)",
    6: "Mayotte",
    11: "Île-de-France",
    24: "Centre-Val de Loire",
    27: "Bourgogne-Franche-Comté",
    28: "Normandie",
    32: "Hauts-de-France",
    44: "Grand Est",
    52: "Pays de la Loire",
    53: "Bretagne",
    75: "Nouvelle-Aquitaine",
    76: "Occitanie",
    84: "Auvergne-Rhône-Alpes",
    93: "Provence-Alpes-Côte d’Azur",
    94: "Corse",
    99: "Non renseignée",
}


def _duckdb_path(path: Path) -> str:
    return path.as_posix().replace("'", "''")


def _csp_sources() -> list[Path]:
    """Return the optimized CSP Parquet source(s) available to DuckDB.

    A consolidated ``csp_core.parquet`` (or ``csp_core_multiyear.parquet``)
    takes precedence.  Otherwise all year-level ``csp_core_*.parquet`` files
    are read with DuckDB's ``union_by_name`` support.  The explicit
    ``CSP_DATA_PATH`` environment variable always wins, which keeps existing
    deployments and tests deterministic.
    """
    if CSP_DATA_ENV:
        return [CSP_PATH] if CSP_PATH.exists() else []
    for candidate_name in ("csp_core.parquet", "csp_core_multiyear.parquet"):
        candidate = CSP_DIR / candidate_name
        if candidate.exists():
            return [candidate]
    candidates = sorted(CSP_DIR.rglob("csp_core_*.parquet"))
    return candidates


def _mortality_rows(path: Path) -> list[tuple[str, str, int, str, int, float | None, bool]]:
    """Normalize the CépiDc workbook's wide blocks to rows for DuckDB."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[1]
    groups = {
        "ensemble": "ensemble", "femmes": "femmes", "hommes": "hommes",
        "0-64 ans": "0-64", "65-84 ans": "65-84", "85 ans ou plus": "85+",
    }
    column_specs: list[tuple[int, str, int]] = []
    current_group: str | None = None
    for column in range(2, sheet.max_column + 1):
        block = sheet.cell(1, column).value
        if block is not None:
            current_group = groups.get(str(block).strip().lower())
        year = sheet.cell(2, column).value
        # The workbook stores the year headers as text (e.g. ``"2015"``),
        # although some Excel writers may emit numeric cells.  Normalize both
        # representations so the loader does not silently drop every column.
        try:
            year_int = int(str(year).strip()) if year is not None else None
        except (TypeError, ValueError):
            year_int = None
        if current_group and year_int is not None:
            column_specs.append((column, current_group, year_int))
    output: list[tuple[str, str, int, str, int, float | None]] = []
    cause_order = 0
    for row in range(3, sheet.max_row + 1):
        label = sheet.cell(row, 1).value
        if label is None or str(label).strip() == "":
            continue
        cause_label = str(label).strip()
        code = f"c{cause_order:03d}"
        is_detail = cause_label.casefold().startswith("dont ")
        for column, group, year in column_specs:
            value = sheet.cell(row, column).value
            deaths = float(value) if isinstance(value, (int, float)) else None
            output.append((code, cause_label, cause_order, group, year, deaths, is_detail))
        cause_order += 1
    workbook.close()
    # Keep the order column in the physical table for a stable metadata list.
    return output


class DamirRepository:
    """Small read-only query layer shared by the prototype endpoints."""

    def __init__(self) -> None:
        missing = [path.name for path in (CUBE_PATH, TRANSCO_PATH) if not path.exists()]
        if missing:
            raise RuntimeError(f"Fichier(s) DAMIR introuvable(s) : {', '.join(missing)}")

        self._cursor_lock = threading.Lock()
        self._thread_local = threading.local()
        self._connection = duckdb.connect(database=":memory:")
        thread_count = max(2, min(os.cpu_count() or 4, 8))
        self._connection.execute(f"SET threads = {thread_count}")
        self._connection.execute("SET enable_object_cache = true")
        self._connection.execute("SET preserve_insertion_order = false")
        self._connection.execute(
            f"CREATE VIEW cube AS SELECT * FROM read_parquet('{_duckdb_path(CUBE_PATH)}')"
        )
        self._connection.execute(
            f"""
            CREATE TABLE transco AS
            SELECT
                TRY_CAST(PRS_NAT AS BIGINT) AS prs_nat,
                libelle,
                grand_poste,
                poste,
                sous_poste
            FROM read_csv_auto(
                '{_duckdb_path(TRANSCO_PATH)}',
                delim = ';',
                header = true,
                all_varchar = true
            )
            WHERE TRY_CAST(PRS_NAT AS BIGINT) IS NOT NULL
            """
        )
        self._connection.execute("ANALYZE transco")
        self.has_delays = DELAYS_PATH.exists()
        if self.has_delays:
            self._connection.execute(
                f"CREATE VIEW delays AS SELECT * FROM read_parquet('{_duckdb_path(DELAYS_PATH)}')"
            )
        self.has_pathologies = PATHOLOGIES_PATH.exists()
        if self.has_pathologies:
            self._connection.execute(
                f"CREATE VIEW pathologies AS SELECT * FROM read_parquet('{_duckdb_path(PATHOLOGIES_PATH)}')"
            )
        self.csp_paths = _csp_sources()
        self.has_csp = bool(self.csp_paths)
        self.csp_file_size = sum(path.stat().st_size for path in self.csp_paths) if self.has_csp else 0
        if self.has_csp:
            paths_sql = ", ".join(f"'{_duckdb_path(path)}'" for path in self.csp_paths)
            # All optimized year files use the same logical schema.  UNION BY
            # NAME also keeps the view readable while a new year is being
            # added with a slightly different physical column order.
            self._connection.execute(
                f"CREATE VIEW csp AS SELECT * FROM read_parquet([{paths_sql}], union_by_name = true)"
            )
        self.has_mortality = MORTALITY_PATH.exists()
        self.mortality_file_size = MORTALITY_PATH.stat().st_size if self.has_mortality else 0
        if self.has_mortality:
            if MORTALITY_PATH.suffix.lower() == ".parquet":
                self._connection.execute(
                    f"""CREATE VIEW mortality AS
                    SELECT cause_code, cause_label, cause_order,
                           CASE sexe_code
                             WHEN 'age_0_64' THEN '0-64'
                             WHEN 'age_65_84' THEN '65-84'
                             WHEN 'age_85_plus' THEN '85+'
                             ELSE sexe_code
                           END AS population_group,
                           annee AS year,
                           effectif_deces AS deaths, is_detail
                    FROM read_parquet('{_duckdb_path(MORTALITY_PATH)}')"""
                )
            else:
                rows = _mortality_rows(MORTALITY_PATH)
                self._connection.execute(
                    """CREATE TABLE mortality (
                        cause_code VARCHAR, cause_label VARCHAR, cause_order INTEGER,
                        population_group VARCHAR, year INTEGER, deaths DOUBLE,
                        is_detail BOOLEAN DEFAULT false
                    )"""
                )
                self._connection.executemany(
                    "INSERT INTO mortality (cause_code, cause_label, cause_order, population_group, year, deaths, is_detail) VALUES (?, ?, ?, ?, ?, ?, ?)", rows
                )
            # Parquet-backed sources are views and cannot be ANALYZEd; DuckDB
            # already uses the Parquet statistics for these read-only queries.
            if MORTALITY_PATH.suffix.lower() != ".parquet":
                self._connection.execute("ANALYZE mortality")

    def query(self, sql: str, params: list[Any] | None = None) -> list[dict[str, Any]]:
        connection = getattr(self._thread_local, "connection", None)
        if connection is None:
            # DuckDB recommande un curseur/connexion par thread. Les vues Parquet et la
            # table de correspondance restent partagées, tandis que les requêtes HTTP
            # en lecture ne sont plus sérialisées par un verrou global.
            with self._cursor_lock:
                connection = self._connection.cursor()
            self._thread_local.connection = connection
        cursor = connection.execute(sql, params or [])
        columns = [item[0] for item in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]

    @lru_cache(maxsize=1)
    def metadata(self) -> dict[str, Any]:
        annual = self.query(
            """
            SELECT soi_ann AS year, SUM(rem)::DOUBLE AS reimbursed
            FROM cube
            WHERE soi_ann >= 2015
            GROUP BY 1
            HAVING SUM(rem) > 0
            ORDER BY 1
            """
        )
        if not annual:
            raise RuntimeError("Aucune année exploitable dans le cube DAMIR")

        maximum = max(float(row["reimbursed"] or 0) for row in annual)
        years = [int(row["year"]) for row in annual if float(row["reimbursed"] or 0) >= maximum * 0.01]

        grand_posts = self.query(
            """
            SELECT COALESCE(t.grand_poste, 'Autres') AS label,
                   SUM(c.rem)::DOUBLE AS amount
            FROM cube c
            LEFT JOIN transco t USING (prs_nat)
            GROUP BY 1
            ORDER BY 2 DESC
            """
        )
        region_codes = self.query(
            "SELECT DISTINCT TRY_CAST(region AS INTEGER) AS code FROM cube WHERE region IS NOT NULL ORDER BY 1"
        )
        regions = [
            {"code": int(row["code"]), "label": REGIONS.get(int(row["code"]), f"Région {row['code']}")}
            for row in region_codes
            if row["code"] is not None
        ]
        default_start_year = 2015 if 2015 in years else min(years)
        default_end_year = 2024 if 2024 in years else max(years)
        return {
            "years": years,
            "default_start_year": default_start_year,
            "default_end_year": default_end_year,
            "grand_posts": [str(row["label"]) for row in grand_posts],
            "regions": regions,
            "source": "Open DAMIR · Assurance Maladie",
            "cube_size_bytes": CUBE_PATH.stat().st_size,
        }

    @staticmethod
    def _filters(
        start_year: int,
        end_year: int,
        grand_poste: str | None,
        region: int | None,
    ) -> tuple[str, list[Any]]:
        clauses = ["c.soi_ann BETWEEN ? AND ?"]
        params: list[Any] = [start_year, end_year]
        if grand_poste:
            clauses.append("COALESCE(t.grand_poste, 'Autres') = ?")
            params.append(grand_poste)
        if region is not None:
            clauses.append("c.region = ?")
            params.append(region)
        return " AND ".join(clauses), params

    def dashboard(
        self,
        start_year: int,
        end_year: int,
        grand_poste: str | None,
        region: int | None,
    ) -> dict[str, Any]:
        where, params = self._filters(start_year, end_year, grand_poste, region)
        annual = self.query(
            f"""
            SELECT
                c.soi_ann AS year,
                SUM(c.rem)::DOUBLE AS reimbursed,
                SUM(c.dep)::DOUBLE AS expense,
                SUM(c.qte)::DOUBLE AS quantity,
                (100.0 * SUM(c.rem) / NULLIF(SUM(c.dep), 0))::DOUBLE AS coverage_rate
            FROM cube c
            LEFT JOIN transco t USING (prs_nat)
            WHERE {where}
            GROUP BY 1
            ORDER BY 1
            """,
            params,
        )
        if not annual:
            return {
                "summary": None,
                "annual": [],
                "breakdown": [],
                "top_services": [],
                "scope": {},
            }

        total_reimbursed = sum(float(row["reimbursed"] or 0) for row in annual)
        total_expense = sum(float(row["expense"] or 0) for row in annual)
        total_quantity = sum(float(row["quantity"] or 0) for row in annual)
        first = float(annual[0]["reimbursed"] or 0)
        last = float(annual[-1]["reimbursed"] or 0)
        change_pct = (100 * (last / first - 1)) if first else None

        dimension = "COALESCE(t.poste, 'Non classé')" if grand_poste else "COALESCE(t.grand_poste, 'Autres')"
        breakdown_label = "Postes" if grand_poste else "Grands postes"
        breakdown = self.query(
            f"""
            SELECT {dimension} AS label, SUM(c.rem)::DOUBLE AS amount
            FROM cube c
            LEFT JOIN transco t USING (prs_nat)
            WHERE {where}
            GROUP BY 1
            HAVING SUM(c.rem) > 0
            ORDER BY 2 DESC
            LIMIT 7
            """,
            params,
        )
        if (
            grand_poste
            and breakdown
            and total_reimbursed
            and float(breakdown[0]["amount"] or 0) / total_reimbursed >= 0.95
        ):
            breakdown = self.query(
                f"""
                SELECT COALESCE(t.sous_poste, 'Non classé') AS label,
                       SUM(c.rem)::DOUBLE AS amount
                FROM cube c
                LEFT JOIN transco t USING (prs_nat)
                WHERE {where}
                GROUP BY 1
                HAVING SUM(c.rem) > 0
                ORDER BY 2 DESC
                LIMIT 7
                """,
                params,
            )
            breakdown_label = "Sous-postes"
        top_services = self.query(
            f"""
            SELECT
                c.prs_nat AS code,
                COALESCE(ANY_VALUE(t.libelle), 'Prestation non libellée') AS label,
                SUM(c.rem)::DOUBLE AS amount
            FROM cube c
            LEFT JOIN transco t USING (prs_nat)
            WHERE {where}
            GROUP BY c.prs_nat
            HAVING SUM(c.rem) > 0
            ORDER BY 3 DESC
            LIMIT 8
            """,
            params,
        )

        return {
            "summary": {
                "reimbursed": total_reimbursed,
                "expense": total_expense,
                "quantity": total_quantity,
                "coverage_rate": 100 * total_reimbursed / total_expense if total_expense else None,
                "change_pct": change_pct,
            },
            "annual": annual,
            "breakdown": breakdown,
            "top_services": top_services,
            "scope": {
                "start_year": start_year,
                "end_year": end_year,
                "grand_poste": grand_poste,
                "region": region,
                "breakdown_label": breakdown_label,
            },
        }

    def export_rows(
        self,
        start_year: int,
        end_year: int,
        grand_poste: str | None,
        region: int | None,
    ) -> list[dict[str, Any]]:
        where, params = self._filters(start_year, end_year, grand_poste, region)
        return self.query(
            f"""
            SELECT
                c.soi_ann AS annee_soins,
                COALESCE(t.grand_poste, 'Autres') AS grand_poste,
                c.region AS code_region,
                SUM(c.rem)::DOUBLE AS montant_rembourse,
                SUM(c.dep)::DOUBLE AS depense_totale,
                SUM(c.qte)::DOUBLE AS quantite_actes,
                (100.0 * SUM(c.rem) / NULLIF(SUM(c.dep), 0))::DOUBLE AS taux_prise_en_charge
            FROM cube c
            LEFT JOIN transco t USING (prs_nat)
            WHERE {where}
            GROUP BY 1, 2, 3
            ORDER BY 1, 4 DESC
            """,
            params,
        )


repository = DamirRepository()

app = FastAPI(
    title="DAMIR Studio API",
    description="API locale de consultation du cube Open DAMIR.",
    version="0.1.0",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=False,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)


@app.middleware("http")
async def frontend_cache_control(request: Request, call_next):
    response = await call_next(request)
    content_type = response.headers.get("content-type", "")
    if request.url.path == "/" or content_type.startswith("text/html"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        response.headers["Pragma"] = "no-cache"
    elif request.url.path.startswith("/assets/") and "cache-control" not in response.headers:
        response.headers["Cache-Control"] = "public, max-age=31536000, immutable"
    return response


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/meta")
@lru_cache(maxsize=1)
def metadata() -> dict[str, Any]:
    base = repository.metadata()
    studio = studio_metadata(repository)
    base["default_start_year"] = 2015 if 2015 in base["years"] else min(base["years"])
    base["default_end_year"] = 2024 if 2024 in base["years"] else max(base["years"])
    return {**base, **analysis_metadata(repository, REGIONS), **studio}


@app.get("/api/options")
def options(
    grand_post: str | None = None,
    post: str | None = None,
    sub_post: str | None = None,
) -> dict[str, Any]:
    return hierarchy_options(repository, grand_post, post, sub_post)


@app.post("/api/analysis")
def analysis(payload: AnalysisRequest) -> dict[str, Any]:
    try:
        return run_analysis(repository, payload, REGIONS)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.post("/api/workbench")
def workbench(payload: WorkbenchRequest) -> dict[str, Any]:
    try:
        return run_workbench(repository, payload, REGIONS)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.get("/api/panorama")
def panorama_view(
    start_year: int = Query(..., ge=2000, le=2100),
    end_year: int = Query(..., ge=2000, le=2100),
    reference_year: int | None = Query(default=None, ge=2000, le=2100),
    measure: str = Query(default="reimbursed"),
    grand_post: str | None = None,
    region: int | None = None,
) -> dict[str, Any]:
    try:
        return panorama(repository, start_year, end_year, grand_post, region, reference_year, measure)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.get("/api/methodology")
def methodology_view() -> dict[str, Any]:
    return methodology(repository)


@app.get("/api/pathologies/meta")
def pathologies_metadata_view() -> dict[str, Any]:
    try:
        return pathology_metadata(repository)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.post("/api/pathologies/overview")
def pathologies_overview_view(payload: PathologyOverviewRequest) -> dict[str, Any]:
    try:
        return pathology_overview(repository, payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.get("/api/csp/meta")
def csp_metadata_view() -> dict[str, Any]:
    try:
        return csp_metadata(repository)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.post("/api/csp/overview")
def csp_overview_view(payload: CspOverviewRequest) -> dict[str, Any]:
    try:
        return csp_overview(repository, payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.post("/api/csp/evolution")
def csp_evolution_view(payload: CspEvolutionRequest) -> dict[str, Any]:
    try:
        return csp_evolution(repository, payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.get("/api/mortality/meta")
def mortality_metadata_view() -> dict[str, Any]:
    try:
        return mortality_metadata(repository)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.post("/api/mortality/overview")
def mortality_overview_view(payload: MortalityOverviewRequest) -> dict[str, Any]:
    try:
        return mortality_overview(repository, payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.post("/api/mortality/extraction/preview")
def mortality_extraction_preview_view(payload: MortalityExtractionRequest) -> dict[str, Any]:
    try:
        return mortality_extraction_preview(repository, payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.post("/api/mortality/extraction.csv")
def mortality_extraction_csv(payload: MortalityExtractionRequest) -> StreamingResponse:
    try:
        rows = mortality_extraction_rows(repository, payload)
        columns = mortality_extraction_columns(payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";")
    writer.writerow([column["label"] for column in columns])
    for row in rows:
        writer.writerow([row.get(column["key"]) for column in columns])
    headers = {"Content-Disposition": 'attachment; filename="mortalite_extraction.csv"'}
    return StreamingResponse(iter(["\ufeff" + output.getvalue()]), media_type="text/csv; charset=utf-8",
                             headers=headers)


@app.post("/api/mortality/extraction.xlsx")
def mortality_extraction_xlsx(payload: MortalityExtractionRequest) -> StreamingResponse:
    try:
        rows = mortality_extraction_rows(repository, payload)
        columns = mortality_extraction_columns(payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    metadata = mortality_metadata(repository)
    selected_cause = next((item for item in metadata["causes"] if item["code"] == payload.cause), None)
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Données")
    sheet.freeze_panes = "A2"
    column_widths = {
        "year": 12,
        "cause": 48,
        "population": 24,
        "deaths": 18,
        "share": 20,
    }
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = column_widths.get(column["key"], 22)
    header_fill = PatternFill(fill_type="solid", fgColor="18243A")
    header_cells = []
    for column in columns:
        cell = WriteOnlyCell(sheet, value=column["label"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        header_cells.append(cell)
    sheet.append(header_cells)
    for row in rows:
        cells = []
        for column in columns:
            cell = WriteOnlyCell(sheet, value=row.get(column["key"]))
            if column["kind"] == "quantity":
                cell.number_format = "#,##0"
            elif column["kind"] == "percent":
                cell.number_format = "0.00"
            cells.append(cell)
        sheet.append(cells)
    metadata_sheet = workbook.create_sheet("Métadonnées")
    metadata_sheet.freeze_panes = "A2"
    metadata_sheet.column_dimensions["A"].width = 24
    metadata_sheet.column_dimensions["B"].width = 110
    metadata_rows = [
        ["Élément", "Valeur"],
        ["Source", metadata["source"]],
        ["Date d’extraction", datetime.now().astimezone().isoformat(timespec="seconds")],
        ["Champ", metadata["scope"]],
        ["Période", f"{payload.start_year}–{payload.end_year}"],
        ["Cause", selected_cause["label"] if selected_cause else "Toutes les causes disponibles"],
        ["Population", MORTALITY_GROUPS[payload.population] if payload.population in MORTALITY_GROUPS else "Tous les périmètres publiés"],
        ["Dimensions", ", ".join(MORTALITY_DIMENSIONS[key][0] for key in payload.dimensions)],
        ["Mesures", ", ".join(MORTALITY_MEASURES[key][0] for key in payload.measures)],
        ["Précaution", "Effectifs bruts non rapportés à une population exposée ; aucune comparaison de risque ne doit être déduite."],
        ["Valeurs absentes", "Une cellule non disponible ou non applicable reste vide et n’est jamais remplacée par zéro."],
    ]
    for row_index, values in enumerate(metadata_rows):
        cells = []
        for value in values:
            cell = WriteOnlyCell(metadata_sheet, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index == 0:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            elif len(cells) == 0:
                cell.font = Font(bold=True, color="18243A")
            cells.append(cell)
        metadata_sheet.append(cells)
    buffer = io.BytesIO()
    workbook.save(buffer)
    headers = {"Content-Disposition": 'attachment; filename="mortalite_extraction.xlsx"'}
    return StreamingResponse(iter([buffer.getvalue()]),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)


@app.get("/api/csp/regions.geojson")
def csp_regions_geojson() -> FileResponse:
    if not CSP_GEOJSON_PATH.exists():
        raise HTTPException(status_code=404, detail="Le fond de carte régional n’est pas disponible.")
    return FileResponse(CSP_GEOJSON_PATH, media_type="application/geo+json")


@app.post("/api/csp/extraction/preview")
def csp_extraction_preview_view(payload: CspExtractionRequest) -> dict[str, Any]:
    try:
        return csp_extraction_preview(repository, payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.post("/api/csp/extraction.csv")
def csp_extraction_csv(payload: CspExtractionRequest) -> StreamingResponse:
    try:
        rows = csp_extraction_rows(repository, payload)
        columns = csp_extraction_columns(payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";")
    writer.writerow([column["label"] for column in columns])
    for row in rows:
        writer.writerow([row.get(column["key"]) for column in columns])
    headers = {"Content-Disposition": f'attachment; filename="csp_extraction_{payload.year}.csv"'}
    return StreamingResponse(iter(["\ufeff" + output.getvalue()]), media_type="text/csv; charset=utf-8",
                             headers=headers)


@app.post("/api/csp/extraction.xlsx")
def csp_extraction_xlsx(payload: CspExtractionRequest) -> StreamingResponse:
    try:
        rows = csp_extraction_rows(repository, payload)
        columns = csp_extraction_columns(payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    metadata = csp_metadata(repository)
    selected_level = next(item for item in metadata["levels"] if item["key"] == payload.level)
    selected_csp = next((item for item in selected_level["options"] if item["code"] == payload.csp_code), None)
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Données")
    sheet.append([column["label"] for column in columns])
    for row in rows:
        cells = []
        for column in columns:
            cell = WriteOnlyCell(sheet, value=row.get(column["key"]))
            if column["kind"] == "quantity":
                cell.number_format = "#,##0"
            elif column["kind"] == "percent":
                cell.number_format = "0.00"
            cells.append(cell)
        sheet.append(cells)
    metadata_sheet = workbook.create_sheet("Métadonnées")
    metadata_sheet.append(["Élément", "Valeur"])
    metadata_sheet.append(["Source", metadata.get("source", "Recensement de la population · Insee")])
    metadata_sheet.append(["Date d’extraction", datetime.now().astimezone().isoformat(timespec="seconds")])
    metadata_sheet.append(["Millésime", payload.year])
    metadata_sheet.append(["Champ", "Actifs ayant un emploi (TACT = 11)"])
    metadata_sheet.append(["Niveau CSP", selected_level["label"]])
    metadata_sheet.append(["CSP", selected_csp["label"] if selected_csp else payload.csp_code])
    metadata_sheet.append(["Dimensions", ", ".join(CSP_DIMENSIONS[key][0] for key in payload.dimensions)])
    metadata_sheet.append(["Mesures", ", ".join(CSP_MEASURES[key][0] for key in payload.measures)])
    metadata_sheet.append(["Pondération", "Effectifs calculés avec le poids individuel IPONDI"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    headers = {"Content-Disposition": f'attachment; filename="csp_extraction_{payload.year}.xlsx"'}
    return StreamingResponse(iter([buffer.getvalue()]),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)


@app.post("/api/pathologies/extraction/preview")
def pathologies_extraction_preview_view(payload: PathologyExtractionRequest) -> dict[str, Any]:
    try:
        return pathology_extraction_preview(repository, payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.post("/api/pathologies/extraction.csv")
def pathologies_extraction_csv(payload: PathologyExtractionRequest) -> StreamingResponse:
    try:
        rows = pathology_extraction_rows(repository, payload)
        columns = pathology_extraction_columns(payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";")
    writer.writerow([column["label"] for column in columns])
    for row in rows:
        writer.writerow([row.get(column["key"]) for column in columns])
    headers = {"Content-Disposition": 'attachment; filename="pathologies_extraction.csv"'}
    return StreamingResponse(iter(["\ufeff" + output.getvalue()]), media_type="text/csv; charset=utf-8",
                             headers=headers)


@app.post("/api/pathologies/extraction.xlsx")
def pathologies_extraction_xlsx(payload: PathologyExtractionRequest) -> StreamingResponse:
    try:
        rows = pathology_extraction_rows(repository, payload)
        columns = pathology_extraction_columns(payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Données")
    sheet.append([column["label"] for column in columns])
    for row in rows:
        cells = []
        for column in columns:
            cell = WriteOnlyCell(sheet, value=row.get(column["key"]))
            if column["kind"] == "quantity":
                cell.number_format = "#,##0"
            elif column["kind"] == "percent":
                cell.number_format = "0.0"
            cells.append(cell)
        sheet.append(cells)
    metadata_sheet = workbook.create_sheet("Métadonnées")
    metadata_sheet.append(["Élément", "Valeur"])
    metadata_sheet.append(["Source", "Cartographie des pathologies · Cnam"])
    metadata_sheet.append(["Date d’extraction", datetime.now().astimezone().isoformat(timespec="seconds")])
    metadata_sheet.append(["Code pathologie", payload.top])
    metadata_sheet.append(["Période", f"{payload.start_year}–{payload.end_year}"])
    metadata_sheet.append(["Dimensions", ", ".join(PATHOLOGY_DIMENSIONS[key][0] for key in payload.dimensions)])
    metadata_sheet.append(["Mesures", ", ".join(PATHOLOGY_MEASURES[key][0] for key in payload.measures)])
    metadata_sheet.append(["Secret statistique", "Les effectifs inférieurs à 10 patients peuvent être masqués à la source."])
    buffer = io.BytesIO()
    workbook.save(buffer)
    headers = {"Content-Disposition": 'attachment; filename="pathologies_extraction.xlsx"'}
    return StreamingResponse(iter([buffer.getvalue()]),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)


@app.post("/api/extraction/preview")
def extraction(payload: ExtractionRequest) -> dict[str, Any]:
    try:
        return extraction_preview(repository, payload, REGIONS)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


def _rows_to_csv(rows: list[dict[str, Any]]) -> str:
    output = io.StringIO(newline="")
    if not rows:
        return ""
    writer = csv.DictWriter(output, fieldnames=list(rows[0]), delimiter=";")
    writer.writeheader()
    writer.writerows(rows)
    return "\ufeff" + output.getvalue()


@app.post("/api/extraction.csv")
def extraction_csv(payload: ExtractionRequest) -> StreamingResponse:
    try:
        rows = extraction_rows(repository, payload, REGIONS)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    headers = {"Content-Disposition": 'attachment; filename="damir_extraction_avancee.csv"'}
    return StreamingResponse(
        iter([_rows_to_csv(rows)]),
        media_type="text/csv; charset=utf-8",
        headers=headers,
    )


@app.post("/api/extraction.xlsx")
def extraction_xlsx(payload: ExtractionRequest) -> StreamingResponse:
    try:
        rows = extraction_rows(repository, payload, REGIONS)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Données")
    columns = extraction_columns(payload)
    sheet.append([column["label"] for column in columns])
    for row in rows:
        cells = []
        for column in columns:
            cell = WriteOnlyCell(sheet, value=row.get(column["key"]))
            if column["kind"] in ("money", "quantity"):
                cell.number_format = "#,##0"
            elif column["kind"] == "percent":
                cell.number_format = "0.0"
            cells.append(cell)
        sheet.append(cells)
    metadata_sheet = workbook.create_sheet("Métadonnées")
    metadata_sheet.append(["Élément", "Valeur"])
    metadata_sheet.append(["Source", "Open DAMIR · Assurance Maladie"])
    metadata_sheet.append(["Date d’extraction", datetime.now().astimezone().isoformat(timespec="seconds")])
    metadata_sheet.append(["Période", f"{payload.start_year}–{payload.end_year}"])
    metadata_sheet.append(["Grand poste", payload.grand_post or "Tous"])
    metadata_sheet.append(["Poste", payload.post or "Tous"])
    metadata_sheet.append(["Sous-poste", payload.sub_post or "Tous"])
    metadata_sheet.append(["Prestations", ", ".join(map(str, payload.service_codes)) or "Toutes"])
    metadata_sheet.append(["Sexes", ", ".join(map(str, payload.sexes)) or "Tous"])
    metadata_sheet.append(["Âges", ", ".join(map(str, payload.ages)) or "Tous"])
    metadata_sheet.append(["Régions", ", ".join(map(str, payload.regions)) or "Toutes"])
    metadata_sheet.append(["Assurances", ", ".join(map(str, payload.insurances)) or "Toutes"])
    metadata_sheet.append(["Enveloppes", ", ".join(map(str, payload.envelopes)) or "Toutes"])
    metadata_sheet.append(["ALD", "Toutes" if payload.ald is None else "ALD" if payload.ald == 1 else "Hors ALD"])
    metadata_sheet.append(["Dimensions", ", ".join(DIMENSIONS[key][0] for key in payload.dimensions)])
    metadata_sheet.append(["Mesures", ", ".join(METRICS[key].label for key in payload.measures)])
    metadata_sheet.append([])
    metadata_sheet.append(["Mesure", "Définition", "Formule", "Précaution"])
    definitions = methodology(repository)["measures"]
    selected = {item["key"]: item for item in definitions if item["key"] in payload.measures}
    for key in payload.measures:
        item = selected[key]
        metadata_sheet.append([item["label"], item["definition"], item["formula"], item["caveat"] or ""])
    reliability = methodology(repository)["reliability"]
    metadata_sheet.append([])
    metadata_sheet.append(["Consolidation", reliability.get("status", "Indisponible")])
    buffer = io.BytesIO()
    workbook.save(buffer)
    headers = {"Content-Disposition": 'attachment; filename="damir_extraction_avancee.xlsx"'}
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _validated_period(start_year: int, end_year: int) -> tuple[int, int]:
    if start_year > end_year:
        raise HTTPException(status_code=422, detail="La période sélectionnée est invalide.")
    return start_year, end_year


@app.get("/api/dashboard")
def dashboard(
    start_year: int = Query(..., ge=2000, le=2100),
    end_year: int = Query(..., ge=2000, le=2100),
    grand_poste: str | None = None,
    region: int | None = None,
) -> dict[str, Any]:
    _validated_period(start_year, end_year)
    return repository.dashboard(start_year, end_year, grand_poste, region)


@app.get("/api/export.csv")
def export_csv(
    start_year: int = Query(..., ge=2000, le=2100),
    end_year: int = Query(..., ge=2000, le=2100),
    grand_poste: str | None = None,
    region: int | None = None,
) -> StreamingResponse:
    _validated_period(start_year, end_year)
    rows = repository.export_rows(start_year, end_year, grand_poste, region)
    output = io.StringIO(newline="")
    fieldnames = [
        "annee_soins",
        "grand_poste",
        "code_region",
        "region",
        "montant_rembourse",
        "depense_totale",
        "quantite_actes",
        "taux_prise_en_charge",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=";")
    writer.writeheader()
    for row in rows:
        code = row.get("code_region")
        writer.writerow(
            {
                **row,
                "region": REGIONS.get(int(code), f"Région {code}") if code is not None else "Non renseignée",
            }
        )
    payload = "\ufeff" + output.getvalue()
    headers = {"Content-Disposition": 'attachment; filename="damir_extraction.csv"'}
    return StreamingResponse(iter([payload]), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/assets/{asset_name}")
def frontend_asset(asset_name: str) -> FileResponse:
    """Serve hashed assets and recover transparently from a stale HTML bundle."""
    if Path(asset_name).name != asset_name:
        raise HTTPException(status_code=404, detail="Fichier d’interface introuvable.")
    requested = FRONTEND_ASSETS / asset_name
    if requested.is_file():
        return FileResponse(
            requested,
            headers={"Cache-Control": "public, max-age=31536000, immutable"},
        )

    suffix = Path(asset_name).suffix
    logical_names = (
        "index",
        "PanoramaPage",
        "PathologyPage",
        "CspPage",
        "MortalityPage",
        "AnalysisPage",
        "ExtractionPage",
        "MethodologyPage",
        "AdvancedFilterPanel",
        "MultiSelect",
        "utils",
        "vendor-react",
        "vendor-plotly",
    )
    logical_name = next(
        (name for name in logical_names if asset_name.startswith(f"{name}-")),
        None,
    )
    if logical_name:
        candidates = sorted(
            FRONTEND_ASSETS.glob(f"{logical_name}-*{suffix}"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        if candidates:
            return FileResponse(
                candidates[0],
                headers={
                    "Cache-Control": "no-store, no-cache, must-revalidate",
                    "X-DAMIR-Asset-Recovered": "1",
                },
            )
    raise HTTPException(status_code=404, detail="Fichier d’interface introuvable.")


if FRONTEND_DIST.exists():
    app.mount("/", StaticFiles(directory=FRONTEND_DIST, html=True), name="frontend")
