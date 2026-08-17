"""La fabrique d'exports, écrite une fois.

Le motif CSV + Excel était recopié **cinq fois** dans `main.py` — DAMIR,
Pathologies, CSP, Population, Mortalité — pour plus de 400 lignes quasi
identiques. Le principe du dépôt, extraire au troisième usage réel, était
atteint et dépassé.

Ce qui variait vraiment d'une source à l'autre tient en cinq champs, et ce sont
les cinq champs de `ExportSpec` : le nom de fichier, les colonnes, les lignes,
les largeurs, et le contenu de la feuille Métadonnées. Tout le reste — le
point-virgule, la nomenclature d'encodage, les en-têtes figées, le remplissage
du bandeau, les formats de nombre, la feuille Métadonnées elle-même — était le
même partout, à des détails près qui étaient des oublis plutôt que des choix.

**Ce que la mise en commun fait gagner**, et qui n'est donc pas une perte :
DAMIR, CSP et Pathologies n'avaient ni en-têtes figées, ni largeurs de colonne,
ni bandeau coloré ; ils les ont maintenant. Aucune capacité ne disparaît.

**Ce qui reste réglable par source**, parce que c'en est vraiment un choix :
le format des pourcentages — DAMIR et Pathologies affichent une décimale, les
trois autres deux — et les blocs libres ajoutés sous les métadonnées, dont se
sert le dictionnaire des mesures DAMIR.

La limite de 250 000 lignes n'est pas ici : elle vit dans le
`*_extraction_rows` de chaque source, où elle protège la requête et pas
seulement le fichier.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: Le bleu d'encre du bandeau d'en-tête. Il ne vient pas de `theme.css` : un
#: classeur s'ouvre dans Excel, hors de l'application, et n'a pas de thème.
HEADER_FILL = "18243A"

#: Largeur par défaut d'une colonne dont la source ne dit rien.
DEFAULT_WIDTH = 22


@dataclass
class ExportSpec:
    """Tout ce qui distingue l'export d'une source de celui d'une autre."""

    #: Nom du fichier, sans extension : `population_extraction`, `csp_extraction_2023`…
    filename: str
    #: `{key, label, kind}` — `kind` pilote le format de nombre.
    columns: list[dict[str, Any]]
    rows: list[dict[str, Any]]
    #: Les lignes de la feuille Métadonnées, en-tête compris.
    metadata: list[list[Any]]
    #: Largeur par clé de colonne. Les clés absentes prennent `DEFAULT_WIDTH`.
    widths: dict[str, int] = field(default_factory=dict)
    #: DAMIR et Pathologies affichent une décimale, les autres deux. C'est un
    #: choix de source, pas un oubli : un taux de prise en charge se lit à
    #: 0,1 point près, une part de population à 0,01.
    percent_format: str = "0.00"
    #: Blocs libres ajoutés sous les métadonnées, séparés par une ligne vide.
    #: DAMIR y range le dictionnaire des mesures et l'état de consolidation.
    extra_blocks: list[list[list[Any]]] = field(default_factory=list)


def _stamp() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def metadata_header(source: str) -> list[list[Any]]:
    """Les trois lignes que toute feuille Métadonnées porte.

    La date d'extraction en fait partie : une donnée consolidée depuis n'aurait
    pas les mêmes chiffres, et un classeur qui circule doit dire de quand il
    date.
    """
    return [
        ["Élément", "Valeur"],
        ["Source", source],
        ["Date d’extraction", _stamp()],
    ]


def csv_response(spec: ExportSpec) -> StreamingResponse:
    """Le CSV : point-virgule, et une nomenclature d'octets en tête.

    Le séparateur point-virgule et le BOM ne sont pas des détails : sans eux,
    Excel en configuration française ouvre le fichier en une seule colonne et
    remplace les accents par des losanges. C'est la première chose qu'on nous
    a signalée, et elle vaut pour les cinq sources.
    """
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";")
    writer.writerow([column["label"] for column in spec.columns])
    for row in spec.rows:
        writer.writerow([row.get(column["key"]) for column in spec.columns])
    return StreamingResponse(
        iter(["﻿" + output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{spec.filename}.csv"'},
    )


def _header_cells(sheet: Any, labels: list[Any]) -> list[Any]:
    fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    cells = []
    for label in labels:
        cell = WriteOnlyCell(sheet, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cells.append(cell)
    return cells


def xlsx_response(spec: ExportSpec) -> StreamingResponse:
    """Le classeur auto-documenté : une feuille de données, une de méthode.

    La feuille Métadonnées n'est pas une politesse. Un fichier de chiffres qui
    circule sans son périmètre, sa date et ses précautions finit par être lu de
    travers ; c'est la seule chose qui parte avec lui pour l'en empêcher.
    """
    workbook = Workbook(write_only=True)

    sheet = workbook.create_sheet("Données")
    sheet.freeze_panes = "A2"
    for index, column in enumerate(spec.columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = (
            spec.widths.get(column["key"], DEFAULT_WIDTH))
    sheet.append(_header_cells(sheet, [column["label"] for column in spec.columns]))

    for row in spec.rows:
        cells = []
        for column in spec.columns:
            cell = WriteOnlyCell(sheet, value=row.get(column["key"]))
            if column["kind"] in ("money", "quantity"):
                cell.number_format = "#,##0"
            elif column["kind"] == "percent":
                cell.number_format = spec.percent_format
            cells.append(cell)
        sheet.append(cells)

    metadata_sheet = workbook.create_sheet("Métadonnées")
    metadata_sheet.freeze_panes = "A2"
    metadata_sheet.column_dimensions["A"].width = 24
    metadata_sheet.column_dimensions["B"].width = 110
    # Les blocs libres suivent les métadonnées, séparés par une ligne vide, et
    # leur première ligne porte l'en-tête : c'est ainsi que le dictionnaire des
    # mesures DAMIR se lit comme un tableau et non comme une suite de lignes.
    blocks = [spec.metadata, *spec.extra_blocks]
    for block_index, block in enumerate(blocks):
        if block_index:
            metadata_sheet.append([])
        for row_index, values in enumerate(block):
            if row_index == 0:
                metadata_sheet.append(_header_cells(metadata_sheet, values))
                continue
            cells = []
            for position, value in enumerate(values):
                cell = WriteOnlyCell(metadata_sheet, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if position == 0:
                    cell.font = Font(bold=True, color=HEADER_FILL)
                cells.append(cell)
            metadata_sheet.append(cells)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{spec.filename}.xlsx"'},
    )
