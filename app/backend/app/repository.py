"""L'accès aux données : les chemins, les vues DuckDB, le dépôt.

Cette couche vivait en tête de `main.py`, qui mêlait donc trois choses — où
sont les fichiers, comment on les interroge, et quelles routes on sert. Elle
est isolée depuis le point 2.4, sans rien changer au comportement.

**Tout est en vue, rien n'est chargé en mémoire.** DuckDB lit les Parquet là où
ils sont posés. Une seule exception : la table de transcodage (152 Ko) est
matérialisée, parce qu'elle est jointe à presque chaque requête.

**Une base absente ne fait pas tomber l'application.** Seuls le cube et la
table de transcodage sont obligatoires ; le reste renseigne un drapeau
`has_*`, et l'écran correspondant disparaît proprement.
"""

from __future__ import annotations

import os
import threading
from functools import lru_cache
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook

from .cache import DiskCache, fingerprint


APP_DIR = Path(__file__).resolve().parent
BACKEND_DIR = APP_DIR.parent
WEBAPP_DIR = BACKEND_DIR.parent
PROJECT_DIR = WEBAPP_DIR.parent
DATA_DIR = Path(os.environ.get("DAMIR_DATA_DIR", PROJECT_DIR / "data")).resolve()
CUBE_PATH = DATA_DIR / "cube_damir.parquet"
# Cube agrégé à l'année par `tools/build_cube_compact.py`. Huit fois plus petit
# que le cube brut et strictement équivalent pour toutes les requêtes de
# l'application, qui n'utilisent jamais le mois de soins. Optionnel : sans lui,
# le cube brut est lu directement, simplement plus lentement.
COMPACT_CUBE_PATH = DATA_DIR / "cube_damir_compact.parquet"
DELAYS_PATH = DATA_DIR / "cube_delais.parquet"
TRANSCO_PATH = DATA_DIR / "prs_nat_transco.csv"
CACHE_DIR = DATA_DIR / ".cache"
PATHOLOGIES_PATH = Path(os.environ.get(
    "PATHOLOGIES_DATA_PATH", DATA_DIR / "pathologies" / "effectifs.parquet"
)).resolve()
CSP_DATA_ENV = os.environ.get("CSP_DATA_PATH")
CSP_DIR = (DATA_DIR / "csp").resolve()
CSP_PATH = Path(CSP_DATA_ENV).resolve() if CSP_DATA_ENV else (CSP_DIR / "csp_core.parquet").resolve()
CSP_GEOJSON_PATH = Path(os.environ.get(
    "CSP_GEOJSON_PATH", DATA_DIR / "pathologies" / "regions.geojson"
)).resolve()
POPULATION_DIR = (DATA_DIR / "population").resolve()
# Produit par `tools/build_population.py` à partir du classeur Insee. Optionnel :
# sans lui, l'application fonctionne — la base Population disparaît et les
# mesures par habitant retombent sur la population de référence de la Cnam, en
# le disant.
POPULATION_PATH = Path(os.environ.get(
    "POPULATION_DATA_PATH", POPULATION_DIR / "population.parquet"
)).resolve()
MORTALITY_DATA_ENV = os.environ.get("MORTALITY_DATA_PATH")
MORTALITY_DIR = (DATA_DIR / "mortalite").resolve()
MORTALITY_XLSX_PATH = MORTALITY_DIR / "Effectifs_cause_mortalite.xlsx"
MORTALITY_PARQUET_PATH = MORTALITY_DIR / "mortalite_core.parquet"
# The compact Parquet is the runtime source.  The workbook remains a safe
# fallback so a fresh checkout can still start before the conversion step.
MORTALITY_PATH = Path(MORTALITY_DATA_ENV).resolve() if MORTALITY_DATA_ENV else (
    MORTALITY_PARQUET_PATH if MORTALITY_PARQUET_PATH.exists() else MORTALITY_XLSX_PATH
).resolve()
FRONTEND_DIST = WEBAPP_DIR / "frontend" / "dist"
FRONTEND_ASSETS = FRONTEND_DIST / "assets"

REGIONS = {
    1: "Guadeloupe",
    2: "Martinique",
    3: "Guyane",
    4: "La Réunion",
    5: "DOM (ancien code)",
    6: "Mayotte",
    11: "Île-de-France",
    24: "Centre-Val de Loire",
    27: "Bourgogne-Franche-Comté",
    28: "Normandie",
    32: "Hauts-de-France",
    44: "Grand Est",
    52: "Pays de la Loire",
    53: "Bretagne",
    75: "Nouvelle-Aquitaine",
    76: "Occitanie",
    84: "Auvergne-Rhône-Alpes",
    93: "Provence-Alpes-Côte d’Azur",
    94: "Corse",
    99: "Non renseignée",
}


def _duckdb_path(path: Path) -> str:
    return path.as_posix().replace("'", "''")


def _csp_sources() -> list[Path]:
    """Return the optimized CSP Parquet source(s) available to DuckDB.

    A consolidated ``csp_core.parquet`` (or ``csp_core_multiyear.parquet``)
    takes precedence.  Otherwise all year-level ``csp_core_*.parquet`` files
    are read with DuckDB's ``union_by_name`` support.  The explicit
    ``CSP_DATA_PATH`` environment variable always wins, which keeps existing
    deployments and tests deterministic.
    """
    if CSP_DATA_ENV:
        return [CSP_PATH] if CSP_PATH.exists() else []
    for candidate_name in ("csp_core.parquet", "csp_core_multiyear.parquet"):
        candidate = CSP_DIR / candidate_name
        if candidate.exists():
            return [candidate]
    candidates = sorted(CSP_DIR.rglob("csp_core_*.parquet"))
    return candidates


def _mortality_rows(path: Path) -> list[tuple[str, str, int, str, int, float | None, bool]]:
    """Normalize the CépiDc workbook's wide blocks to rows for DuckDB."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[1]
    groups = {
        "ensemble": "ensemble", "femmes": "femmes", "hommes": "hommes",
        "0-64 ans": "0-64", "65-84 ans": "65-84", "85 ans ou plus": "85+",
    }
    column_specs: list[tuple[int, str, int]] = []
    current_group: str | None = None
    for column in range(2, sheet.max_column + 1):
        block = sheet.cell(1, column).value
        if block is not None:
            current_group = groups.get(str(block).strip().lower())
        year = sheet.cell(2, column).value
        # The workbook stores the year headers as text (e.g. ``"2015"``),
        # although some Excel writers may emit numeric cells.  Normalize both
        # representations so the loader does not silently drop every column.
        try:
            year_int = int(str(year).strip()) if year is not None else None
        except (TypeError, ValueError):
            year_int = None
        if current_group and year_int is not None:
            column_specs.append((column, current_group, year_int))
    output: list[tuple[str, str, int, str, int, float | None]] = []
    cause_order = 0
    for row in range(3, sheet.max_row + 1):
        label = sheet.cell(row, 1).value
        if label is None or str(label).strip() == "":
            continue
        cause_label = str(label).strip()
        code = f"c{cause_order:03d}"
        is_detail = cause_label.casefold().startswith("dont ")
        for column, group, year in column_specs:
            value = sheet.cell(row, column).value
            deaths = float(value) if isinstance(value, (int, float)) else None
            output.append((code, cause_label, cause_order, group, year, deaths, is_detail))
        cause_order += 1
    workbook.close()
    # Keep the order column in the physical table for a stable metadata list.
    return output


class DamirRepository:
    """Small read-only query layer shared by the prototype endpoints."""

    def __init__(self) -> None:
        missing = [path.name for path in (CUBE_PATH, TRANSCO_PATH) if not path.exists()]
        if missing:
            raise RuntimeError(f"Fichier(s) DAMIR introuvable(s) : {', '.join(missing)}")

        self._cursor_lock = threading.Lock()
        self._thread_local = threading.local()
        self._connection = duckdb.connect(database=":memory:")
        thread_count = max(2, min(os.cpu_count() or 4, 8))
        self._connection.execute(f"SET threads = {thread_count}")
        self._connection.execute("SET enable_object_cache = true")
        self._connection.execute("SET preserve_insertion_order = false")
        self.cube_path = self._resolve_cube()
        self._connection.execute(
            f"CREATE VIEW cube AS SELECT * FROM read_parquet('{_duckdb_path(self.cube_path)}')"
        )
        self._connection.execute(
            f"""
            CREATE TABLE transco AS
            SELECT
                TRY_CAST(PRS_NAT AS BIGINT) AS prs_nat,
                libelle,
                grand_poste,
                poste,
                sous_poste
            FROM read_csv_auto(
                '{_duckdb_path(TRANSCO_PATH)}',
                delim = ';',
                header = true,
                all_varchar = true
            )
            WHERE TRY_CAST(PRS_NAT AS BIGINT) IS NOT NULL
            """
        )
        self._connection.execute("ANALYZE transco")
        self.has_delays = DELAYS_PATH.exists()
        if self.has_delays:
            self._connection.execute(
                f"CREATE VIEW delays AS SELECT * FROM read_parquet('{_duckdb_path(DELAYS_PATH)}')"
            )
        self.has_pathologies = PATHOLOGIES_PATH.exists()
        if self.has_pathologies:
            self._connection.execute(
                f"CREATE VIEW pathologies AS SELECT * FROM read_parquet('{_duckdb_path(PATHOLOGIES_PATH)}')"
            )
        self.csp_paths = _csp_sources()
        self.has_csp = bool(self.csp_paths)
        self.csp_file_size = sum(path.stat().st_size for path in self.csp_paths) if self.has_csp else 0
        if self.has_csp:
            paths_sql = ", ".join(f"'{_duckdb_path(path)}'" for path in self.csp_paths)
            # All optimized year files use the same logical schema.  UNION BY
            # NAME also keeps the view readable while a new year is being
            # added with a slightly different physical column order.
            self._connection.execute(
                f"CREATE VIEW csp AS SELECT * FROM read_parquet([{paths_sql}], union_by_name = true)"
            )
        self.has_population = POPULATION_PATH.exists()
        self.population_file_size = POPULATION_PATH.stat().st_size if self.has_population else 0
        if self.has_population:
            self._connection.execute(
                f"CREATE VIEW population AS SELECT * FROM read_parquet('{_duckdb_path(POPULATION_PATH)}')"
            )
        self.has_mortality = MORTALITY_PATH.exists()
        self.mortality_file_size = MORTALITY_PATH.stat().st_size if self.has_mortality else 0
        if self.has_mortality:
            if MORTALITY_PATH.suffix.lower() == ".parquet":
                self._connection.execute(
                    f"""CREATE VIEW mortality AS
                    SELECT cause_code, cause_label, cause_order,
                           CASE sexe_code
                             WHEN 'age_0_64' THEN '0-64'
                             WHEN 'age_65_84' THEN '65-84'
                             WHEN 'age_85_plus' THEN '85+'
                             ELSE sexe_code
                           END AS population_group,
                           annee AS year,
                           effectif_deces AS deaths, is_detail
                    FROM read_parquet('{_duckdb_path(MORTALITY_PATH)}')"""
                )
            else:
                rows = _mortality_rows(MORTALITY_PATH)
                self._connection.execute(
                    """CREATE TABLE mortality (
                        cause_code VARCHAR, cause_label VARCHAR, cause_order INTEGER,
                        population_group VARCHAR, year INTEGER, deaths DOUBLE,
                        is_detail BOOLEAN DEFAULT false
                    )"""
                )
                self._connection.executemany(
                    "INSERT INTO mortality (cause_code, cause_label, cause_order, population_group, year, deaths, is_detail) VALUES (?, ?, ?, ?, ?, ?, ?)", rows
                )
            # Parquet-backed sources are views and cannot be ANALYZEd; DuckDB
            # already uses the Parquet statistics for these read-only queries.
            if MORTALITY_PATH.suffix.lower() != ".parquet":
                self._connection.execute("ANALYZE mortality")

    @staticmethod
    def _resolve_cube() -> Path:
        """Choisit le cube compact quand il est à jour, le cube brut sinon.

        Le cube compact est un dérivé : s'il est plus ancien que sa source, il
        décrit des données périmées. Mieux vaut alors être lent et juste, en
        signalant qu'une reconstruction est à faire.
        """
        if not COMPACT_CUBE_PATH.exists():
            return CUBE_PATH
        if COMPACT_CUBE_PATH.stat().st_mtime < CUBE_PATH.stat().st_mtime:
            print(
                f"[DAMIR] {COMPACT_CUBE_PATH.name} est plus ancien que {CUBE_PATH.name} : "
                "lecture du cube brut. Relancez « python tools/build_cube_compact.py »."
            )
            return CUBE_PATH
        return COMPACT_CUBE_PATH

    def query(self, sql: str, params: list[Any] | None = None) -> list[dict[str, Any]]:
        connection = getattr(self._thread_local, "connection", None)
        if connection is None:
            # DuckDB recommande un curseur/connexion par thread. Les vues Parquet et la
            # table de correspondance restent partagées, tandis que les requêtes HTTP
            # en lecture ne sont plus sérialisées par un verrou global.
            with self._cursor_lock:
                connection = self._connection.cursor()
            self._thread_local.connection = connection
        cursor = connection.execute(sql, params or [])
        columns = [item[0] for item in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]

    @lru_cache(maxsize=1)
    def metadata(self) -> dict[str, Any]:
        annual = self.query(
            """
            SELECT soi_ann AS year, SUM(rem)::DOUBLE AS reimbursed
            FROM cube
            WHERE soi_ann >= 2015
            GROUP BY 1
            HAVING SUM(rem) > 0
            ORDER BY 1
            """
        )
        if not annual:
            raise RuntimeError("Aucune année exploitable dans le cube DAMIR")

        maximum = max(float(row["reimbursed"] or 0) for row in annual)
        years = [int(row["year"]) for row in annual if float(row["reimbursed"] or 0) >= maximum * 0.01]

        grand_posts = self.query(
            """
            SELECT COALESCE(t.grand_poste, 'Autres') AS label,
                   SUM(c.rem)::DOUBLE AS amount
            FROM cube c
            LEFT JOIN transco t USING (prs_nat)
            GROUP BY 1
            ORDER BY 2 DESC
            """
        )
        region_codes = self.query(
            "SELECT DISTINCT TRY_CAST(region AS INTEGER) AS code FROM cube WHERE region IS NOT NULL ORDER BY 1"
        )
        regions = [
            {"code": int(row["code"]), "label": REGIONS.get(int(row["code"]), f"Région {row['code']}")}
            for row in region_codes
            if row["code"] is not None
        ]
        default_start_year = 2015 if 2015 in years else min(years)
        default_end_year = 2024 if 2024 in years else max(years)
        return {
            "years": years,
            "default_start_year": default_start_year,
            "default_end_year": default_end_year,
            "grand_posts": [str(row["label"]) for row in grand_posts],
            "regions": regions,
            "source": "Open DAMIR · Assurance Maladie",
            "cube_size_bytes": CUBE_PATH.stat().st_size,
        }


repository = DamirRepository()
