from __future__ import annotations

import asyncio
import unittest
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

from openpyxl import load_workbook

from app.analysis import ExtractionRequest, FilterPayload, extraction_preview
from app.csp import CspExtractionRequest, CspOverviewRequest, csp_extraction_preview, csp_metadata, csp_overview
from app.main import REGIONS, csp_extraction_xlsx, extraction_xlsx, pathologies_extraction_xlsx, repository
from app.pathologies import (
    PathologyExtractionRequest,
    PathologyOverviewRequest,
    pathology_extraction_preview,
    pathology_metadata,
    pathology_overview,
)


class StudioTests(unittest.TestCase):
    def assert_amount_equal(self, actual: float, expected: float) -> None:
        self.assertAlmostEqual(actual, expected, delta=max(abs(expected) * 1e-12, 1e-6))

    def test_repository_supports_parallel_readers(self) -> None:
        def total_for_year(year: int) -> float:
            row = repository.query(
                "SELECT SUM(rem)::DOUBLE AS value FROM cube WHERE soi_ann = ?",
                [year],
            )[0]
            return float(row["value"] or 0)

        years = [2021, 2022, 2023, 2024]
        with ThreadPoolExecutor(max_workers=4) as executor:
            parallel = list(executor.map(total_for_year, years))
        sequential = [total_for_year(year) for year in years]

        for actual, expected in zip(parallel, sequential):
            self.assert_amount_equal(actual, expected)

    # Les onze tests du moteur « Reperes » sont partis avec lui (point 2.1).
    # Ce qu'ils verrouillaient de precieux — qu'un taux moyen n'est pas la
    # moyenne des taux annuels — est desormais verrouille cote Tableau par
    # `tests/test_pivot.py`, qui verifie que les composantes sont sommees
    # avant l'application de la formule.

    def test_excel_export_contains_data_and_metadata_sheets(self) -> None:
        response = extraction_xlsx(ExtractionRequest(
            start_year=2024,
            end_year=2024,
            dimensions=["year"],
            measures=["reimbursed"],
        ))

        async def response_bytes() -> bytes:
            return b"".join([chunk async for chunk in response.body_iterator])

        workbook = load_workbook(BytesIO(asyncio.run(response_bytes())), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Données", "Métadonnées"])
        self.assertEqual(workbook["Données"]["A1"].value, "Année de soins")
        self.assertEqual(workbook["Données"]["B1"].value, "Montant remboursé")
        self.assertEqual(workbook["Données"]["B2"].number_format, "#,##0")
        self.assertEqual(workbook["Métadonnées"]["A1"].value, "Élément")
        self.assertEqual(workbook["Métadonnées"]["A2"].value, "Source")
        metadata_values = {row[0].value: row[1].value for row in workbook["Métadonnées"].iter_rows()
                           if len(row) >= 2 and row[0].value}
        self.assertEqual(metadata_values["Dimensions"], "Année de soins")
        self.assertEqual(metadata_values["Mesures"], "Montant remboursé")

    def test_pathology_overview_exposes_identity_profiles_and_territories(self) -> None:
        metadata = pathology_metadata(repository)
        diabetes = next(
            pathology
            for family in metadata["families"]
            for group in family["groups"]
            for pathology in [{"code": group["code"], "label": group["label"]}, *group["pathologies"]]
            if "diab" in pathology["label"].lower()
        )
        overview = pathology_overview(repository, PathologyOverviewRequest(
            top=diabetes["code"],
            year=metadata["default_year"],
        ))

        self.assertEqual(len(overview["kpis"]), 4)
        self.assertGreater(len(overview["annual"]), 1)
        self.assertGreater(len(overview["age_sex"]), 1)
        self.assertGreater(len(overview["territories"]), 1)
        self.assertGreater(overview["kpis"][0]["value"], 0)
        self.assertIn("femme pour 1 homme", overview["kpis"][3]["detail"])
        self.assertNotIn(",0 femme", overview["kpis"][3]["detail"])
        self.assertEqual(metadata["levels"], 3)
        self.assertTrue(any(group["pathologies"] for family in metadata["families"] for group in family["groups"]))

    def test_pathology_overview_applies_population_filters(self) -> None:
        overview = pathology_overview(repository, PathologyOverviewRequest(
            top="DIA_CAT_CAT",
            year=2024,
            region="11",
            age="60-64",
            sex="femmes",
        ))
        direct = repository.query(
            """SELECT ntop::DOUBLE AS patients FROM pathologies
               WHERE top = 'DIA_CAT_CAT' AND year(annee) = 2024 AND dept = '999'
                 AND region = '11' AND cla_age_5 = '60-64' AND libelle_sexe = 'femmes'"""
        )[0]

        self.assert_amount_equal(overview["kpis"][0]["value"], float(direct["patients"]))
        self.assertEqual(overview["context"]["region"], "11")
        self.assertEqual(overview["context"]["age"], "60-64")
        self.assertEqual(overview["context"]["sex"], "femmes")

    def test_masked_pathology_prevalence_is_never_replaced_by_zero(self) -> None:
        overview = pathology_overview(repository, PathologyOverviewRequest(
            top="TAA_CAT_EXC",
            year=2024,
            age="85-89",
            sex="tous sexes",
        ))
        missing = [item for item in overview["territories"] if item["prevalence"] is None]

        self.assertTrue(missing)
        self.assertEqual(overview["quality"]["unavailable_territories"], len(missing))
        self.assertTrue(all(item["patients"] is None for item in missing))

    def test_pathology_extraction_uses_requested_dimensions_and_measures(self) -> None:
        metadata = pathology_metadata(repository)
        selected = metadata["families"][0]
        preview = pathology_extraction_preview(repository, PathologyExtractionRequest(
            top=selected["code"],
            start_year=metadata["default_year"],
            end_year=metadata["default_year"],
            dimensions=["year", "sex", "region"],
            measures=["patients", "prevalence"],
            limit=100,
        ))

        self.assertEqual(
            [column["key"] for column in preview["columns"]],
            ["year", "sex", "region", "patients", "prevalence"],
        )
        self.assertGreater(preview["total_rows"], 0)
        self.assertTrue(preview["rows"])

    def test_average_measure_is_available_in_full_scope_extractions(self) -> None:
        preview = extraction_preview(repository, ExtractionRequest(
            start_year=2024,
            end_year=2024,
            dimensions=["year"],
            measures=["average_reimbursed"],
        ), REGIONS)
        direct = repository.query(
            """SELECT SUM(rem)::DOUBLE / NULLIF(SUM(qte), 0) AS value
               FROM cube WHERE soi_ann = 2024"""
        )[0]

        self.assert_amount_equal(float(preview["rows"][0]["average_reimbursed"]), float(direct["value"]))

    def test_pathology_excel_export_keeps_numeric_formats_and_metadata(self) -> None:
        metadata = pathology_metadata(repository)
        selected = metadata["families"][0]
        response = pathologies_extraction_xlsx(PathologyExtractionRequest(
            top=selected["code"],
            start_year=metadata["default_year"],
            end_year=metadata["default_year"],
            dimensions=["year", "region"],
            measures=["patients", "prevalence"],
        ))

        async def response_bytes() -> bytes:
            return b"".join([chunk async for chunk in response.body_iterator])

        workbook = load_workbook(BytesIO(asyncio.run(response_bytes())), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Données", "Métadonnées"])
        self.assertEqual(workbook["Données"]["C2"].number_format, "#,##0")
        self.assertEqual(workbook["Données"]["D2"].number_format, "0.0")
        self.assertEqual(workbook["Métadonnées"]["A2"].value, "Source")
        metadata_values = {row[0].value: row[1].value for row in workbook["Métadonnées"].iter_rows()
                           if len(row) >= 2 and row[0].value}
        self.assertEqual(metadata_values["Dimensions"], "Année, Région")
        self.assertEqual(metadata_values["Mesures"], "Patients, Prévalence")

    def test_csp_metadata_exposes_both_levels_and_all_regions(self) -> None:
        metadata = csp_metadata(repository)

        levels = {item["key"]: item for item in metadata["levels"]}
        self.assertEqual(metadata["years"], list(range(2015, 2024)))
        self.assertEqual(len(levels["groupe_6"]["options"]), 6)
        self.assertEqual(len(levels["categorie_29"]["options"]), 29)
        self.assertEqual(len(metadata["regions"]), 18)  # France + 17 régions
        self.assertLess(metadata["core_size_bytes"], 15_000_000)

    def test_csp_overview_reconciles_with_the_optimized_parquet(self) -> None:
        result = csp_overview(repository, CspOverviewRequest(
            level="groupe_6", csp_code="3", region="FR", age="all", sex=0,
        ))
        direct = repository.query(
            """SELECT SUM(effectif)::DOUBLE AS population,
                      SUM(CASE WHEN code_csp = '3' THEN effectif ELSE 0 END)::DOUBLE AS selected
                 FROM csp WHERE annee = 2023 AND niveau_csp = 'groupe_6'"""
        )[0]

        self.assert_amount_equal(result["kpis"][0]["value"], float(direct["population"]))
        self.assert_amount_equal(result["kpis"][1]["value"], float(direct["selected"]))
        self.assertEqual(len(result["territories"]), 17)
        self.assertEqual(len(result["composition"]), 6)
        self.assertEqual(len(result["age_sex"]), 16)

    def test_csp_population_filters_change_the_denominator(self) -> None:
        metadata = csp_metadata(repository)
        age = metadata["ages"][3]["code"]
        result = csp_overview(repository, CspOverviewRequest(
            level="categorie_29", csp_code="38", region="11", age=age, sex=2,
        ))
        direct = repository.query(
            """SELECT SUM(effectif)::DOUBLE AS population,
                      SUM(CASE WHEN code_csp = '38' THEN effectif ELSE 0 END)::DOUBLE AS selected
                 FROM csp
                WHERE annee = 2023 AND niveau_csp = 'categorie_29'
                  AND code_region = '11' AND tranche_age = ? AND code_sexe = 2""",
            [age],
        )[0]

        self.assert_amount_equal(result["kpis"][0]["value"], float(direct["population"]))
        self.assert_amount_equal(result["kpis"][1]["value"], float(direct["selected"]))
        self.assertEqual(result["context"]["region"], "11")
        self.assertEqual(result["context"]["sex"], 2)

    def test_csp_extraction_and_excel_are_documented(self) -> None:
        request = CspExtractionRequest(
            level="categorie_29", csp_code="38",
            dimensions=["region", "age", "sex"],
            measures=["effectif", "population", "share"],
        )
        preview = csp_extraction_preview(repository, request)
        self.assertEqual([column["key"] for column in preview["columns"]],
                         ["region", "age", "sex", "effectif", "population", "share"])
        expected_rows = repository.query(
            """SELECT COUNT(*) AS value FROM (
                   SELECT region, tranche_age, sexe
                     FROM csp
                    WHERE annee = 2023 AND niveau_csp = 'categorie_29'
                    GROUP BY region, tranche_age, sexe
               )"""
        )[0]["value"]
        self.assertEqual(preview["total_rows"], expected_rows)
        self.assertTrue(any(float(row["effectif"]) == 0 for row in preview["rows"]))

        response = csp_extraction_xlsx(request)

        async def response_bytes() -> bytes:
            return b"".join([chunk async for chunk in response.body_iterator])

        workbook = load_workbook(BytesIO(asyncio.run(response_bytes())), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Données", "Métadonnées"])
        self.assertEqual(workbook["Données"]["D2"].number_format, "#,##0")
        self.assertEqual(workbook["Données"]["F2"].number_format, "0.00")
        metadata_values = {row[0].value: row[1].value for row in workbook["Métadonnées"].iter_rows()
                           if len(row) >= 2 and row[0].value}
        self.assertEqual(metadata_values["Champ"], "Actifs ayant un emploi (TACT = 11)")
        self.assertIn("29 catégories", metadata_values["Niveau CSP"])


if __name__ == "__main__":
    unittest.main()
