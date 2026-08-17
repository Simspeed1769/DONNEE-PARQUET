"""Le module d'export, commun aux cinq sources.

Le motif CSV + Excel était recopié cinq fois. La mise en commun ne devait rien
perdre — c'est précisément ce qu'un refactor de ce genre perd en silence, une
source à la fois. Ces tests parcourent **les cinq** et vérifient ce que chaque
copie faisait :

- le CSV s'ouvre dans un Excel français (point-virgule, nomenclature d'octets) ;
- le classeur porte ses deux feuilles, ses en-têtes figées, ses formats ;
- la feuille Métadonnées dit la source, la date et le périmètre ;
- DAMIR garde son dictionnaire des mesures et son état de consolidation.
"""
from __future__ import annotations

import asyncio
import unittest
from io import BytesIO

from openpyxl import load_workbook

from app.analysis import ExtractionRequest
from app.csp import CspExtractionRequest
from app.main import (
    _csp_spec, _damir_spec, _mortality_spec, _pathologies_spec, _population_spec,
    repository,
)
from app.exports import csv_response, xlsx_response
from app.mortality import MortalityExtractionRequest
from app.pathologies import PathologyExtractionRequest
from app.population import PopulationExtractionRequest


def _body(response) -> bytes:
    """Le corps d'une réponse en flux.

    Starlette enveloppe l'itérable synchrone dans un générateur asynchrone :
    il faut donc une boucle d'événements, même pour un fichier déjà en mémoire.
    Même idiome que `tests/test_studio.py`."""
    async def collect() -> bytes:
        return b"".join([
            chunk if isinstance(chunk, bytes) else chunk.encode("utf-8")
            async for chunk in response.body_iterator
        ])
    return asyncio.run(collect())


def _specs() -> list[tuple[str, object]]:
    """Une spécification par source, sur un périmètre volontairement étroit."""
    cases: list[tuple[str, object]] = [
        ("damir", _damir_spec(ExtractionRequest(
            start_year=2024, end_year=2024, dimensions=["year"],
            measures=["reimbursed", "coverage"], limit=50))),
    ]
    if repository.has_pathologies:
        cases.append(("pathologies", _pathologies_spec(PathologyExtractionRequest(
            top="DIA_CAT_CAT", start_year=2022, end_year=2022,
            dimensions=["year"], measures=["patients"], limit=50))))
    if repository.has_csp:
        cases.append(("csp", _csp_spec(CspExtractionRequest(
            year=2022, level="groupe_6", csp_code="3",
            dimensions=["region"], measures=["effectif"], limit=50))))
    if repository.has_population:
        cases.append(("population", _population_spec(PopulationExtractionRequest(
            start_year=2022, end_year=2022, dimensions=["year", "region"],
            measures=["population"], limit=50))))
    if repository.has_mortality:
        cases.append(("mortality", _mortality_spec(MortalityExtractionRequest(
            start_year=2022, end_year=2022, dimensions=["year", "cause", "population"],
            measures=["deaths"], limit=50))))
    return cases


class ExportTests(unittest.TestCase):
    def test_every_source_produces_a_french_excel_csv(self) -> None:
        """Sans point-virgule ni BOM, Excel FR rend une colonne et des losanges."""
        for name, spec in _specs():
            with self.subTest(source=name):
                text = _body(csv_response(spec)).decode("utf-8")
                self.assertTrue(text.startswith("﻿"), "nomenclature d’octets absente")
                header = text.lstrip("﻿").splitlines()[0]
                self.assertIn(";", header)
                # L'en-tête porte les **libellés**, pas les clés techniques.
                self.assertEqual(header.split(";"), [c["label"] for c in spec.columns])

    def test_every_workbook_has_both_sheets_and_frozen_headers(self) -> None:
        for name, spec in _specs():
            with self.subTest(source=name):
                book = load_workbook(BytesIO(_body(xlsx_response(spec))))
                self.assertEqual(book.sheetnames, ["Données", "Métadonnées"])
                self.assertEqual(book["Données"].freeze_panes, "A2")
                self.assertEqual(book["Métadonnées"].freeze_panes, "A2")

    def test_every_metadata_sheet_names_its_source_and_date(self) -> None:
        for name, spec in _specs():
            with self.subTest(source=name):
                sheet = load_workbook(BytesIO(_body(xlsx_response(spec))))["Métadonnées"]
                labels = [row[0].value for row in sheet.iter_rows(min_col=1, max_col=1)]
                self.assertIn("Source", labels)
                self.assertIn("Date d’extraction", labels)

    def test_damir_keeps_its_measure_dictionary_and_consolidation(self) -> None:
        """Le seul export qui embarque un dictionnaire. Il ne devait pas partir
        avec la mise en commun : c'est `extra_blocks` qui le porte."""
        spec = _damir_spec(ExtractionRequest(
            start_year=2024, end_year=2024, dimensions=["year"],
            measures=["reimbursed", "coverage"], limit=50))
        sheet = load_workbook(BytesIO(_body(xlsx_response(spec))))["Métadonnées"]
        cells = [[cell.value for cell in row] for row in sheet.iter_rows()]
        flat = [value for row in cells for value in row]
        self.assertIn("Définition", flat)
        self.assertIn("Formule", flat)
        self.assertIn("Montant remboursé", flat)
        self.assertIn("Taux de prise en charge AMO", flat)
        self.assertIn("Consolidation", flat)

    def test_column_widths_and_number_formats_survive(self) -> None:
        spec = _damir_spec(ExtractionRequest(
            start_year=2024, end_year=2024, dimensions=["year"],
            measures=["reimbursed", "coverage"], limit=50))
        sheet = load_workbook(BytesIO(_body(xlsx_response(spec))))["Données"]
        self.assertEqual(sheet.column_dimensions["A"].width, 12)  # `year`
        formats = {cell.number_format for row in sheet.iter_rows(min_row=2) for cell in row}
        self.assertIn("#,##0", formats)
        # DAMIR garde une décimale sur les pourcentages : c'est un choix de
        # source, conservé par `percent_format`.
        self.assertIn("0.0", formats)

    def test_an_absent_value_stays_empty(self) -> None:
        """La règle du dépôt, vérifiée à la sortie : rien ne remplace un vide
        par zéro, pas même une cellule de tableur."""
        from app.exports import ExportSpec
        spec = ExportSpec(
            filename="essai",
            # Deux colonnes, comme dans la réalité : sur une seule, le module
            # `csv` écrit `""` pour distinguer un champ vide d'une ligne vide,
            # et le test mesurerait cet artefact plutôt que le produit.
            columns=[
                {"key": "annee", "label": "Année", "kind": "raw"},
                {"key": "valeur", "label": "Valeur", "kind": "quantity"},
            ],
            rows=[{"annee": 2024, "valeur": None}],
            metadata=[["Élément", "Valeur"]],
        )
        sheet = load_workbook(BytesIO(_body(xlsx_response(spec))))["Données"]
        self.assertIsNone(sheet.cell(row=2, column=2).value)
        text = _body(csv_response(spec)).decode("utf-8").lstrip("﻿")
        self.assertEqual(text.splitlines()[1], "2024;")


if __name__ == "__main__":
    unittest.main()
