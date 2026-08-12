"""Convertit le classeur Insee des estimations de population en Parquet long.

Source : Insee, *Estimations de population par région, sexe et âge quinquennal*,
`estim-pop-nreg-sexe-aq-1975-2026.xlsx`, mise à jour du 23 décembre 2025.

Ce que le script produit — `data/population/population.parquet`, format long :
`annee`, `region` (code du référentiel de l'application), `region_libelle`,
`sexe`, `age_quinquennal` (code), `age_libelle`, `age_decennal`, `population`.

Trois décisions qui tiennent l'honnêteté du fichier produit :

1. **Le bloc « Ensemble » du classeur n'est pas chargé.** Il est recalculé par
   somme des hommes et des femmes, et le script vérifie que sa propre somme
   retrouve bien le total publié. Charger les trois blocs laisserait le fichier
   se contredire un jour sans que personne ne s'en aperçoive.
2. **Les agrégats ne sont pas des régions.** « France métropolitaine », « DOM »
   et « France métropolitaine et DOM » sont écartés : chargés comme des régions,
   ils compteraient deux fois les mêmes personnes. Le total France se recalcule
   par somme, comme l'Ensemble.
3. **Une donnée absente reste absente.** De 1990 à 1998, l'Insee ne publie pas
   d'âge détaillé au-delà de 90 ans pour la Guadeloupe, la Guyane et la
   Martinique : la case « 95 ans et plus » y est vide et la case « 90 à 94 ans »
   porte en réalité tous les 90 ans et plus. La case vide reste vide, et la
   colonne `age_90_plus_agrege` marque les cellules concernées pour que
   l'application le dise plutôt que de le taire.

Le script est idempotent, réexécutable, et ne modifie jamais le classeur source.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

import duckdb
from openpyxl import load_workbook


# Les blocs de sexe, à leur première colonne. `Ensemble` (colonne B) est lu pour
# contrôle seulement : ce qui est chargé, ce sont les hommes et les femmes.
CHECK_BLOCK = ("Ensemble", 2)
SEX_BLOCKS = [("Hommes", 23), ("Femmes", 44)]
AGE_SPAN = 21  # 20 tranches quinquennales, puis « Total »

HEADER_ROW = 5
FIRST_DATA_ROW = 6

# Les vingt tranches, dans l'ordre des colonnes du classeur. Les codes sont ceux
# que la Cartographie des pathologies emploie déjà : la jointure région × âge ×
# sexe est alors immédiate, sans table de passage.
AGE_BANDS = [
    ("00-04", "0–4 ans", 0), ("05-09", "5–9 ans", 0),
    ("10-14", "10–14 ans", 0), ("15-19", "15–19 ans", 0),
    ("20-24", "20–24 ans", 20), ("25-29", "25–29 ans", 20),
    ("30-34", "30–34 ans", 30), ("35-39", "35–39 ans", 30),
    ("40-44", "40–44 ans", 40), ("45-49", "45–49 ans", 40),
    ("50-54", "50–54 ans", 50), ("55-59", "55–59 ans", 50),
    ("60-64", "60–64 ans", 60), ("65-69", "65–69 ans", 60),
    ("70-74", "70–74 ans", 70), ("75-79", "75–79 ans", 70),
    ("80-84", "80–84 ans", 80), ("85-89", "85–89 ans", 80),
    ("90-94", "90–94 ans", 80), ("95et+", "95 ans et +", 80),
]

# Le référentiel de l'application, par libellé Insee. L'Insee écrit
# « Centre-Val-de-Loire » là où l'application écrit « Centre-Val de Loire », et
# affuble La Réunion d'un astérisque en 1990 : la normalisation se fait ici, une
# fois, plutôt que dans chaque requête.
REGIONS = {
    "auvergne-rhone-alpes": ("84", "Auvergne-Rhône-Alpes"),
    "bourgogne-franche-comte": ("27", "Bourgogne-Franche-Comté"),
    "bretagne": ("53", "Bretagne"),
    "centre-val-de-loire": ("24", "Centre-Val de Loire"),
    "centre-val de loire": ("24", "Centre-Val de Loire"),
    "corse": ("94", "Corse"),
    "grand est": ("44", "Grand Est"),
    "hauts-de-france": ("32", "Hauts-de-France"),
    "ile-de-france": ("11", "Île-de-France"),
    "normandie": ("28", "Normandie"),
    "nouvelle-aquitaine": ("75", "Nouvelle-Aquitaine"),
    "occitanie": ("76", "Occitanie"),
    "pays de la loire": ("52", "Pays de la Loire"),
    "provence-alpes-cote d'azur": ("93", "Provence-Alpes-Côte d’Azur"),
    "provence-alpes-cote d’azur": ("93", "Provence-Alpes-Côte d’Azur"),
    "guadeloupe": ("01", "Guadeloupe"),
    "martinique": ("02", "Martinique"),
    "guyane": ("03", "Guyane"),
    "la reunion": ("04", "La Réunion"),
    "mayotte": ("06", "Mayotte"),
}

# Les lignes de synthèse du classeur : elles ne sont pas des régions, et les
# charger reviendrait à compter deux fois les mêmes personnes.
AGGREGATES = {
    "france metropolitaine",
    "dom",
    "france metropolitaine et dom",
}

# L'âge détaillé au-delà de 90 ans manque sur certaines cellules d'outre-mer des
# années 1990. La note du classeur cite la Guadeloupe, la Guyane et la
# Martinique ; **la Réunion est dans le même cas**, ce que la note ne dit pas.
# Le repère se lit donc dans la donnée — « 95 ans et plus » vide alors que
# « 90 à 94 ans » est renseigné — et non dans une liste écrite à la main, qui
# aurait manqué un territoire sur quatre.


def _fold(value: str) -> str:
    """Le libellé réduit à ce qui l'identifie : sans accent, sans casse, sans
    l'astérisque de note de bas de page que le classeur colle à La Réunion."""
    folded = value.strip().casefold()
    if "(" in folded:
        folded = folded.split("(")[0].strip()
    table = str.maketrans("àâäéèêëîïôöùûüç", "aaaeeeeiioouuuc")
    return folded.translate(table)


def build(source: Path, output: Path) -> dict[str, object]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    rows: list[dict[str, object]] = []
    journal: dict[str, object] = {
        "annees_ingerees": [],
        "annees_ecartees": {},
        "lignes_ecartees": {},
        "libelles_inconnus": {},
        "ecarts_ensemble": [],
        "cellules_90_plus_agregees": 0,
    }

    for sheet_name in workbook.sheetnames:
        if not sheet_name.strip().isdigit():
            # L'onglet « À savoir » porte la notice, pas des données.
            journal["annees_ecartees"][sheet_name.strip()] = "onglet sans données"
            continue
        year = int(sheet_name.strip())
        sheet = workbook[sheet_name]
        # Une feuille ouverte en lecture seule se parcourt d'un bout à l'autre :
        # y piocher cellule par cellule la relit à chaque appel, et les 53
        # onglets y passaient plusieurs minutes.
        grid = [row for row in sheet.iter_rows(values_only=True)]

        def cell(row_number: int, column: int):
            row = grid[row_number - 1] if row_number - 1 < len(grid) else ()
            return row[column - 1] if column - 1 < len(row) else None

        # La mise en page est contrôlée plutôt que supposée : un classeur mis à
        # jour dont les blocs auraient bougé doit se signaler, pas produire des
        # chiffres décalés en silence.
        header = [cell(HEADER_ROW, CHECK_BLOCK[1] + offset) for offset in range(AGE_SPAN)]
        bands_named = all(value and str(value).strip() for value in header[:20])
        total_last = str(header[20]).strip() == "Total"
        blocks_named = all(str(cell(4, column) or "").strip() == name
                           for name, column in [CHECK_BLOCK, *SEX_BLOCKS])
        if not (bands_named and total_last and blocks_named):
            journal["annees_ecartees"][sheet_name] = "mise en page inattendue"
            continue

        ingested = 0
        for row_number in range(FIRST_DATA_ROW, len(grid) + 1):
            raw_label = cell(row_number, 1)
            if raw_label is None or not str(raw_label).strip():
                continue
            label = str(raw_label).strip()
            folded = _fold(label)
            if folded in AGGREGATES:
                journal["lignes_ecartees"].setdefault(label, 0)
                journal["lignes_ecartees"][label] += 1
                continue
            if folded not in REGIONS:
                # Les notes de bas de page vivent dans la colonne A, elles aussi.
                journal["libelles_inconnus"].setdefault(label, 0)
                journal["libelles_inconnus"][label] += 1
                continue

            code, region_label = REGIONS[folded]
            totals = {"Hommes": 0, "Femmes": 0}

            for sex, first_column in SEX_BLOCKS:
                block = [cell(row_number, first_column + offset) for offset in range(20)]
                values = [int(value) if isinstance(value, (int, float)) else None
                          for value in block]
                # « 95 ans et plus » vide alors que « 90 à 94 ans » est
                # renseigné : la dernière case ouverte porte tous les 90 ans et
                # plus, et les deux tranches sont marquées comme telles.
                lumped = values[19] is None and values[18] is not None
                if lumped:
                    journal["cellules_90_plus_agregees"] += 1
                for offset, (band, band_label, decade) in enumerate(AGE_BANDS):
                    population = values[offset]
                    if population is not None:
                        totals[sex] += population
                    rows.append({
                        "annee": year,
                        "region": code,
                        "region_libelle": region_label,
                        "sexe": sex,
                        "age_quinquennal": band,
                        "age_libelle": band_label,
                        "age_decennal": decade,
                        "population": population,
                        # Vraie sur les cellules où « 90 à 94 ans » porte en fait
                        # tous les 90 ans et plus : l'application le dit.
                        "age_90_plus_agrege": bool(lumped and band in {"90-94", "95et+"}),
                    })
            ingested += 1

            # Contrôle : la somme hommes + femmes doit retrouver l'Ensemble
            # publié. Un écart n'est pas corrigé, il est journalisé.
            published = cell(row_number, CHECK_BLOCK[1] + 20)
            if isinstance(published, (int, float)):
                rebuilt = totals["Hommes"] + totals["Femmes"]
                if rebuilt != int(published):
                    journal["ecarts_ensemble"].append({
                        "annee": year, "region": region_label,
                        "recalcule": rebuilt, "publie": int(published),
                        "ecart": rebuilt - int(published),
                    })

        if ingested:
            journal["annees_ingerees"].append(year)
        else:
            journal["annees_ecartees"][sheet_name] = "aucune région reconnue"

    workbook.close()
    if not rows:
        raise ValueError("Aucune ligne exploitable dans le classeur Insee.")

    journal["annees_ingerees"].sort()
    output.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output.with_suffix(".tmp.csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]), delimiter=";")
        writer.writeheader()
        writer.writerows(rows)

    connection = duckdb.connect()
    tmp_output = output.with_suffix(".tmp.parquet")
    tmp_output.unlink(missing_ok=True)
    quoted_csv = csv_path.as_posix().replace("'", "''")
    quoted_out = tmp_output.as_posix().replace("'", "''")
    connection.execute(
        f"COPY (SELECT * FROM read_csv_auto('{quoted_csv}', delim=';', header=true, nullstr='')) "
        f"TO '{quoted_out}' (FORMAT PARQUET, COMPRESSION ZSTD)"
    )
    output.unlink(missing_ok=True)
    tmp_output.replace(output)
    csv_path.unlink(missing_ok=True)

    summary = connection.execute(
        "SELECT count(*) AS lignes, min(annee) AS depuis, max(annee) AS jusqu_a, "
        "       count(DISTINCT region) AS regions, "
        "       count(*) FILTER (WHERE population IS NULL) AS cellules_vides "
        f"FROM read_parquet('{output.as_posix()}')"
    ).fetchone()
    connection.close()

    manifest = {
        "source": str(source),
        "output": str(output),
        "lignes": summary[0],
        "depuis": summary[1],
        "jusqu_a": summary[2],
        "regions": summary[3],
        "cellules_vides": summary[4],
        "journal": journal,
    }
    (output.parent / "population_build_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def _locate_source(root: Path) -> Path:
    """Le classeur, là où il se trouve.

    La feuille de route le place dans `data/population/source/` ; il a été
    déposé dans `data/source/`. On accepte les deux plutôt que d'exiger un
    déplacement de fichier."""
    name = "estim-pop-nreg-sexe-aq-1975-2026.xlsx"
    for candidate in (root / "population" / "source" / name, root / "source" / name):
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        f"Classeur Insee introuvable : déposez « {name} » dans data/population/source/.")


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent / "data"
    result = build(_locate_source(root), root / "population" / "population.parquet")
    print(json.dumps(result, ensure_ascii=False, indent=2))
