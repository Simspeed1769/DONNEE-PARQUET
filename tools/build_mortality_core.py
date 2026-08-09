"""Convert the national CépiDc workbook to a compact long Parquet."""
from __future__ import annotations

import csv
import json
from pathlib import Path

import duckdb
from openpyxl import load_workbook


BLOCKS = [
    ("ensemble", "Tous", 2),
    ("femmes", "Femmes", 12),
    ("hommes", "Hommes", 22),
    ("age_0_64", "0–64 ans", 32),
    ("age_65_84", "65–84 ans", 42),
    ("age_85_plus", "85 ans ou plus", 52),
]


def build(source: Path, output: Path) -> dict[str, object]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["Effectifs"]
    years = [int(sheet.cell(row=2, column=column).value) for column in range(2, 12)]
    rows: list[dict[str, object]] = []
    cause_order = 0
    for row_number in range(3, sheet.max_row + 1):
        label = sheet.cell(row=row_number, column=1).value
        if not label or not str(label).strip():
            continue
        cause_order += 1
        cause_label = str(label).strip()
        code = f"cause_{cause_order:03d}"
        is_detail = cause_label.casefold().startswith("dont ")
        level = 2 if is_detail else 1
        for sex_code, sex_label, first_column in BLOCKS:
            for offset, year in enumerate(years):
                value = sheet.cell(row=row_number, column=first_column + offset).value
                rows.append({
                    "annee": year,
                    "cause_code": code,
                    "cause_label": cause_label,
                    "cause_order": cause_order,
                    "is_detail": is_detail,
                    "niveau": level,
                    "sexe_code": sex_code,
                    "sexe": sex_label,
                    "age_classe": "Tous âges" if sex_code in {"ensemble", "femmes", "hommes"} else sex_label,
                    "effectif_deces": float(value) if value is not None else None,
                    "source": "INSERM–CépiDc, Statistiques nationales sur les causes de décès",
                    "champ": "France entière (Métropole et DROM), résidents décédés en France",
                    "nomenclature": "Liste européenne succincte · CIM selon le millésime",
                })
    output.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output.with_suffix(".tmp.csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]), delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    connection = duckdb.connect()
    tmp_output = output.with_suffix(".tmp.parquet")
    tmp_output.unlink(missing_ok=True)
    connection.execute(
        f"COPY (SELECT * FROM read_csv_auto('{csv_path.as_posix().replace(chr(39), chr(39) * 2)}', delim=';', header=true, nullstr='')) "
        f"TO '{tmp_output.as_posix().replace(chr(39), chr(39) * 2)}' (FORMAT PARQUET, COMPRESSION ZSTD)"
    )
    output.unlink(missing_ok=True)
    tmp_output.replace(output)
    csv_path.unlink(missing_ok=True)
    summary = connection.execute(
        f"SELECT count(*) AS rows, count(DISTINCT annee) AS years, count(DISTINCT cause_code) AS causes FROM read_parquet('{output.as_posix()}')"
    ).fetchone()
    connection.close()
    workbook.close()
    manifest = {"source": str(source), "output": str(output), "rows": summary[0], "years": summary[1], "causes": summary[2]}
    (output.parent / "mortality_build_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


if __name__ == "__main__":
    # Les scripts vivent dans `tools/`, les données dans `data/` : c'est cette
    # dernière que l'on vise, pas le dossier du script.
    root = Path(__file__).resolve().parent.parent / "data" / "mortalite"
    print(json.dumps(build(root / "Effectifs_cause_mortalite.xlsx", root / "mortalite_core.parquet"), ensure_ascii=False, indent=2))
